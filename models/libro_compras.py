from odoo import models, fields, api
from odoo.exceptions import UserError
from datetime import datetime
from dateutil.relativedelta import relativedelta
import io
import csv
import base64


class LibroComprasPeriodo(models.Model):
    _name = 'libro.compras.periodo'
    _description = 'Periodo del Libro de Compras'
    _rec_name = 'periodo'  # Para que el título muestre el periodo

    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        required=True,
        default=lambda self: self.env.company
    )

    incluir_sucursales = fields.Boolean(
        string='Incluir Todas las Sucursales',
        default=False,
        help='Si está marcado, incluirá facturas de todas las sucursales de la empresa'
    )

    # Para parecerse al "Asistente" del módulo antiguo
    assistant_id = fields.Many2one(
        'res.users',
        string='Asistente',
        default=lambda self: self.env.user
    )

    # Puedes dejarlo Char o luego cambiarlo a Many2one si quieres
    contador_name = fields.Char(string='Contador')

    # Fecha de emisión del libro (fecha actual)
    date = fields.Date(string='Fecha Emisión', required=True, default=fields.Date.context_today)

    year = fields.Integer(string='Año', required=True, default=lambda self: fields.Date.context_today(self).year)
    month = fields.Selection([
        ('01', 'Enero'), ('02', 'Febrero'), ('03', 'Marzo'), ('04', 'Abril'),
        ('05', 'Mayo'), ('06', 'Junio'), ('07', 'Julio'), ('08', 'Agosto'),
        ('09', 'Septiembre'), ('10', 'Octubre'), ('11', 'Noviembre'), ('12', 'Diciembre')
    ], string='Mes', required=True)

    # Mes / periodo (texto)
    periodo = fields.Char(
        string='Periodo',
        readonly=True,
        compute='_compute_periodo',
        store=True,
    )

    state = fields.Selection([
        ('draft', 'Borrador'),
        ('validated', 'Validado'),
    ], string='Estado', default='draft')

    invoice_line_ids = fields.One2many(
        'libro.compras.line',
        'periodo_id',
        string='Detalle Compras'
    )

    company_currency_id = fields.Many2one(
        'res.currency',
        related='company_id.currency_id',
        readonly=True,
        store=True,
    )

    # campos calculados
    total_internas_exentas = fields.Monetary(
        string="Total Internas Exentas",
        compute="_compute_totales",
        currency_field="company_currency_id",
    )

    total_internas_gravadas = fields.Monetary(
        string="Total Internas Gravadas",
        compute="_compute_totales",
        currency_field="company_currency_id",
    )

    total_credito_fiscal = fields.Monetary(
        string="Total Crédito Fiscal",
        compute="_compute_totales",
        currency_field="company_currency_id",
    )

    # ----------------- COMPUTADOS -----------------

    @api.depends('year', 'month')
    def _compute_periodo(self):
        for rec in self:
            if rec.year and rec.month:
                month_names = {
                    '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
                    '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
                    '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
                }
                rec.periodo = f"{month_names.get(rec.month, '')} {rec.year}"
            else:
                rec.periodo = ''

    year_display = fields.Char(string='Año (Display)', compute='_compute_year_display', store=False)

    @api.depends('year')
    def _compute_year_display(self):
        for rec in self:
            rec.year_display = str(rec.year) if rec.year else ''

    # ----------------- ACCIONES DE SELECCIÓN -----------------

    def action_select_all(self):
        """Seleccionar todas las líneas."""
        for rec in self:
            rec.invoice_line_ids.write({'select': True})

    def action_unselect_all(self):
        """Deseleccionar todas las líneas."""
        for rec in self:
            rec.invoice_line_ids.write({'select': False})

    # ----------------- RESTRICCIONES -----------------

    def write(self, vals):
        """Bloquea edición si el estado no es 'draft', excepto si se cambia el estado."""
        for rec in self:
            if rec.state != 'draft' and 'state' not in vals:
                raise UserError("Solo puedes modificar libros en estado Borrador.")
        return super().write(vals)

    # ----------------- ACCIONES DE ESTADO -----------------

    _inherit = ['mail.thread', 'mail.activity.mixin']

    state = fields.Selection([
        ('draft', 'Borrador'),
        ('validated', 'Validado'),
    ], string='Estado', default='draft', tracking=True)

    comentarios = fields.Text(string='Comentarios')

    def action_rectify(self):
        """Abre el wizard para rectificar el libro."""
        self.ensure_one()
        return {
            'name': 'Rectificar Libro',
            'type': 'ir.actions.act_window',
            'res_model': 'libro.rectify.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'active_id': self.id, 'active_model': self._name},
        }

    def rectify_book(self, reason):
        """Método llamado por el wizard para ejecutar la rectificación."""
        self.ensure_one()
        self.message_post(body=f"Libro rectificado. Motivo: {reason}", subtype_xmlid="mail.mt_note")
        self.state = 'draft'

    def action_mark_done(self):
        """Equivalente a 'Validar'."""
        for rec in self:
            rec.state = 'validated'

    def action_reset_to_draft(self):
        """Equivalente a 'Cancelar' / volver a borrador."""
        for rec in self:
            rec.state = 'draft'

    # Calculados
    @api.depends("invoice_line_ids.compras_internas_exentas",
                 "invoice_line_ids.compras_internas_gravadas",
                 "invoice_line_ids.credito_fiscal")
    def _compute_totales(self):
        for rec in self:
            rec.total_internas_exentas = sum(rec.invoice_line_ids.mapped("compras_internas_exentas"))
            rec.total_internas_gravadas = sum(rec.invoice_line_ids.mapped("compras_internas_gravadas"))
            rec.total_credito_fiscal = sum(rec.invoice_line_ids.mapped("credito_fiscal"))

    # ----------------- LÓGICA DE LIBRO -----------------

    def action_load_invoices(self):
        """Generar Detalle: carga facturas del mes seleccionado."""
        self.ensure_one()

        if not self.year or not self.month:
            raise UserError("Debe especificar Año y Mes.")

        # Calcular fechas basadas en year/month
        date_from = datetime(self.year, int(self.month), 1).date()
        # Último día del mes
        date_to = (date_from + relativedelta(months=1, days=-1))

        # Determinar las compañías a incluir
        if self.incluir_sucursales:
            # Incluir empresa actual + todas sus sucursales
            company_ids = [self.company_id.id] + self.company_id.child_ids.ids
        else:
            # Solo la empresa actual
            company_ids = [self.company_id.id]

        invoices = self.env['account.move'].search([
            ('move_type', 'in', ['in_invoice', 'in_refund']),  # Incluir facturas Y notas de crédito
            ('state', '=', 'posted'),
            ('invoice_date', '>=', date_from),
            ('invoice_date', '<=', date_to),
            ('company_id', 'in', company_ids),
        ])

        self.invoice_line_ids.unlink()  # limpiar anteriores
        
        valid_count = 0  # Contador de facturas válidas
        skipped_count = 0  # Contador de facturas saltadas
        
        for index, inv in enumerate(invoices, start=1):
            # Extraer información de la factura
            # Buscar código MH en campos personalizados o usar referencia
            codigo_mh = '' # Si tienes campo para esto
            
            # Tipo de documento (puede variar según tu configuración)
            numero_documento = inv.name or ''
            
            # Campos DTE
            numero_control = inv.tgr_l10n_sv_edi_numero_control or ''
            codigo_generacion = inv.tgr_l10n_sv_edi_codigo_generacion or ''
            sello_digital = inv.tgr_l10n_sv_edi_sello_recibido or ''
            
            # EXTRAER TIPO DE DOCUMENTO del nombre/referencia de la factura
            # En facturas de proveedor no existe l10n_latam_document_type_id
            tipo_documento = ''
            ref = inv.ref or inv.name or ''
            
            # Detectar tipo según prefijo en la referencia
            if 'DTE-14' in ref or 'DTE-14' in numero_documento:
                tipo_documento = '14'  # Sujeto Excluido
            elif 'DTE-03' in ref or 'CCF' in ref or 'CCF' in numero_documento:
                tipo_documento = '03'  # Crédito Fiscal
            elif 'DTE-05' in ref or 'NC' in ref:
                tipo_documento = '05'  # Nota de Crédito
            elif 'DTE-06' in ref or 'ND' in ref:
                tipo_documento = '06'  # Nota de Débito
            elif 'DTE-11' in ref:
                tipo_documento = '11'  # Factura Exportación
            elif codigo_generacion:
                # Si tiene código de generación pero no detectamos el tipo, asumir CCF
                tipo_documento = '03'
            else:
                # Por defecto, asumir CCF si no podemos determinar
                tipo_documento = '03'
            
            # DCL (solo para importaciones)
            dcl = ''

            # VALIDAR TIPO DE DOCUMENTO (solo tipos válidos para Hacienda)
            # Según manual oficial, para compras son válidos: 03, 05, 06, 11, 12, 13
            valid_doc_types = ['03', '05', '06', '11', '12', '13']
            if tipo_documento not in valid_doc_types:
                # Saltar documentos con tipo inválido (ej: 14 = Sujeto Excluido)
                skipped_count += 1
                continue

            # Montos: desglosar según tipo de impuesto
            compras_internas_exentas = 0.0
            compras_internas_gravadas = 0.0
            credito_fiscal = 0.0
            
            # IMPORTANTE: Según manual de Hacienda, las notas de crédito (tipo 05)
            # deben reportarse con montos POSITIVOS. El sistema de Hacienda se encarga
            # de restarlas del total automáticamente.
            # No aplicar signo negativo para refunds.

            # Iterar líneas de la factura para calcular montos
            for line in inv.invoice_line_ids:
                amount_line = abs(line.price_subtotal)  # Siempre positivo

                # Determinar si es exento o gravado según impuesto
                if line.tax_ids:
                    # Si tiene impuesto, es gravado
                    compras_internas_gravadas += amount_line
                else:
                    # Si no tiene impuesto, es exento
                    compras_internas_exentas += amount_line

            # IMPORTANTE: Calcular crédito fiscal como exactamente 13% de compras gravadas
            # Esto asegura que cumpla con la validación de Hacienda
            # Nota: Solo compras internas gravadas porque internaciones/importaciones son 0
            credito_fiscal = round(compras_internas_gravadas * 0.13, 2)
            
            # Determinar clase de documento
            clase_doc = '4' if codigo_generacion else '1'
            
            # Total siempre positivo (Hacienda maneja el signo según tipo de documento)
            amount_total = abs(inv.amount_total)
            
            # Usar contador válido para sequence
            valid_count += 1

            self.env['libro.compras.line'].create({
                'periodo_id': self.id,
                'sequence': valid_count,  # Usar contador de válidas
                'move_id': inv.id,
                'partner_id': inv.partner_id.id,
                'invoice_date': inv.invoice_date,
                'codigo_mh': codigo_mh,
                'tipo_documento': tipo_documento,
                'dcl': dcl,
                'numero_documento': numero_documento,
                'numero_control': numero_control,
                'codigo_generacion': codigo_generacion,
                'sello_digital': sello_digital,
                'clase_documento': clase_doc,
                'compras_internas_exentas': compras_internas_exentas,
                'compras_internas_gravadas': compras_internas_gravadas,
                'credito_fiscal': credito_fiscal,
                'amount_total': amount_total,
            })
        
        # Mensaje informativo
        if valid_count == 0 and skipped_count > 0:
            raise UserError(f"No se encontraron facturas válidas para el Libro de Compras.\n"
                          f"Se omitieron {skipped_count} documento(s) con tipo inválido (ej: Sujeto Excluido).\n"
                          f"Tipos válidos para compras: 03, 05, 06, 11, 12, 13")
        
        # Log informativo si hubo documentos omitidos
        if skipped_count > 0:
            import logging
            _logger = logging.getLogger(__name__)
            _logger.warning(f"Libro de Compras: Se cargaron {valid_count} facturas válidas. "
                          f"Se omitieron {skipped_count} documentos con tipo inválido.")
        
        # Forzar recalculo de totales
        self.invalidate_recordset(['invoice_line_ids'])
        
        # Seleccionar automáticamente todas las líneas cargadas
        self.invoice_line_ids.write({'select': True})
        
        # No retornar nada para que Odoo refresque la vista automáticamente

    def _get_document_type(self, invoice):
        """
        Obtener tipo de documento basado en el tipo de movimiento.
        Personaliza según tus necesidades.
        """
        # Por defecto, usar código numérico o descripción
        # Esto es un ejemplo básico
        document_type_map = {
            'in_invoice': '3',  # Factura de compra
            'in_refund': '5',   # Nota de crédito
        }
        return document_type_map.get(invoice.move_type, '3')

        def action_generate_csv(self):
            """Generar Excel (en tu caso CSV) con las facturas seleccionadas."""
            selected = self.invoice_line_ids.filtered(lambda l: l.select)
            if not selected:
                raise UserError("Debe seleccionar al menos una factura.")

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['No', 'Factura', 'Proveedor', 'Fecha', 'Monto'])

            for line in selected:
                writer.writerow([
                    line.sequence or '',
                    line.move_id.name or '',
                    line.partner_id.name or '',
                    line.invoice_date or '',
                    line.amount_total or 0,
                ])

            data = base64.b64encode(output.getvalue().encode())
            output.close()

            attachment = self.env['ir.attachment'].create({
                'name': f'Libro_Compras_{self.periodo or ""}.csv',
                'type': 'binary',
                'datas': data,
                'res_model': self._name,
                'res_id': self.id,
                'mimetype': 'text/csv'
            })
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }

    def action_generate_excel(self):
        """Generar archivo Excel (.xlsx) con las facturas seleccionadas."""
        selected = self.invoice_line_ids.filtered(lambda l: l.select)
        if not selected:
            raise UserError("Debe seleccionar al menos una factura.")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise UserError("La librería 'openpyxl' no está instalada. Instálela con: pip install openpyxl")

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro de Compras"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Encabezados
        headers = [
            'No', 'Fecha Emisión', 'Código MH', 'Tipo de Documento', 'DCL',
            'Número de Documento', 'Número de Control', 'Código de Generación',
            'Sello Digital', 'Proveedor', 'Internas Exentas', 'Internas Gravadas',
            'Crédito Fiscal', 'Total'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row, line in enumerate(selected, start=2):
            ws.cell(row=row, column=1, value=line.sequence or '')
            ws.cell(row=row, column=2, value=str(line.invoice_date) if line.invoice_date else '')
            ws.cell(row=row, column=3, value=line.codigo_mh or '')
            ws.cell(row=row, column=4, value=line.tipo_documento or '')
            ws.cell(row=row, column=5, value=line.dcl or '')
            ws.cell(row=row, column=6, value=line.numero_documento or '')
            ws.cell(row=row, column=7, value=line.numero_control or '')
            ws.cell(row=row, column=8, value=line.codigo_generacion or '')
            ws.cell(row=row, column=9, value=line.sello_digital or '')
            ws.cell(row=row, column=10, value=line.partner_id.name or '')
            ws.cell(row=row, column=11, value=line.compras_internas_exentas or 0)
            ws.cell(row=row, column=12, value=line.compras_internas_gravadas or 0)
            ws.cell(row=row, column=13, value=line.credito_fiscal or 0)
            ws.cell(row=row, column=14, value=line.amount_total or 0)

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        excel_data = base64.b64encode(output.getvalue())
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': f'Libro_Compras_{self.periodo or ""}.xlsx',
            'type': 'binary',
            'datas': excel_data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_generate_csv(self):
        """Generar CSV formato oficial Hacienda (21 columnas, sin encabezados)."""
        selected = self.invoice_line_ids.filtered(lambda l: l.select)
        if not selected:
            raise UserError("Debe seleccionar al menos una factura.")

        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')  # Separador punto y coma
        
        # SIN ENCABEZADOS según manual oficial
        
        # Filas de datos (21 columnas)
        for line in selected:
            # A - Fecha Emisión (DD/MM/YYYY)
            fecha_str = line.invoice_date.strftime('%d/%m/%Y') if line.invoice_date else ''
            
            # B - Clase de Documento (1-4)
            clase_doc = line.clase_documento or '4'
            
            # C - Tipo de Documento (2 caracteres)
            tipo_doc = line.tipo_documento or '03'
            
            # D - Número de Documento (Código Generación sin guiones para DTE)
            if line.codigo_generacion:
                numero_doc = line.codigo_generacion.replace('-', '')
            elif line.numero_control:
                numero_doc = line.numero_control.replace('-', '')
            else:
                numero_doc = line.numero_documento or ''
            
            # E - NIT o NRC del Proveedor (sin guiones)
            nit_nrc = (line.partner_nit or '').replace('-', '')
            
            # F - Nombre del Proveedor
            nombre_prov = line.partner_id.name or ''
            
            # G - Compras Internas Exentas
            g_compras_exentas = f"{line.compras_internas_exentas:.2f}"
            
            # H - Internaciones Exentas
            h_intern_exentas = f"{line.internaciones_exentas:.2f}"
            
            # I - Importaciones Exentas
            i_import_exentas = f"{line.importaciones_exentas:.2f}"
            
            # J - Compras Internas Gravadas
            j_compras_gravadas = f"{line.compras_internas_gravadas:.2f}"
            
            # K - Internaciones Gravadas de Bienes
            k_intern_gravadas = f"{line.internaciones_gravadas_bienes:.2f}"
            
            # L - Importaciones Gravadas de Bienes
            l_import_gravadas_bienes = f"{line.importaciones_gravadas_bienes:.2f}"
            
            # M - Importaciones Gravadas de Servicios
            m_import_gravadas_serv = f"{line.importaciones_gravadas_servicios:.2f}"
            
            # N - Crédito Fiscal
            n_credito_fiscal = f"{line.credito_fiscal:.2f}"
            
            # O - Total de Compras
            o_total = f"{line.amount_total:.2f}"
            
            # P - DUI del Proveedor (9 dígitos, opcional)
            p_dui = (line.dui_proveedor or '').replace('-', '')
            
            # Q - Tipo de Operación
            q_tipo_op = line.tipo_operacion or '1'
            
            # R - Clasificación
            r_clasif = line.clasificacion or '2'
            
            # S - Sector
            s_sector = line.sector or '4'
            
            # T - Tipo Costo/Gasto
            t_tipo_costo = line.tipo_costo_gasto or '5'
            
            # U - Número de Anexo (siempre 3 para compras)
            u_anexo = '3'
            
            writer.writerow([
                fecha_str,           # A
                clase_doc,           # B
                tipo_doc,            # C
                numero_doc,          # D
                nit_nrc,             # E
                nombre_prov,         # F
                g_compras_exentas,   # G
                h_intern_exentas,    # H
                i_import_exentas,    # I
                j_compras_gravadas,  # J
                k_intern_gravadas,   # K
                l_import_gravadas_bienes,  # L
                m_import_gravadas_serv,    # M
                n_credito_fiscal,    # N
                o_total,             # O
                p_dui,               # P
                q_tipo_op,           # Q
                r_clasif,            # R
                s_sector,            # S
                t_tipo_costo,        # T
                u_anexo,             # U
            ])

        data = base64.b64encode(output.getvalue().encode('utf-8'))
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': f'Libro_Compras_Hacienda_{self.periodo or ""}.csv',
            'type': 'binary',
            'datas': data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'text/csv'
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_print_report(self):
        """Imprimir: genera el PDF del libro."""
        return self.env.ref('libros_fiscales.report_libro_compras').report_action(self)
