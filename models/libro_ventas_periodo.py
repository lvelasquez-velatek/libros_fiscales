from odoo import models, fields, api
from odoo.exceptions import UserError
from datetime import datetime
from dateutil.relativedelta import relativedelta
import io
import csv
import base64


class LibroVentasPeriodo(models.Model):
    _name = 'libro.ventas.periodo'
    _description = 'Periodo del Libro de Ventas'
    _rec_name = 'periodo'

    company_id = fields.Many2one(
        'res.company',
        string='Compañía',
        required=True,
        default=lambda self: self.env.company
    )

    incluir_sucursales = fields.Boolean(
        string='Incluir Todas las Sucursales',
        default=False,
        help='Si está marcado, incluirá facturas de todas las sucursales de la empresa'
    )

    assistant_id = fields.Many2one(
        'res.users',
        string='Asistente',
        default=lambda self: self.env.user
    )

    contador_name = fields.Char(string='Contador')

    date = fields.Date(string='Fecha Emisión', required=True, default=fields.Date.context_today)

    year = fields.Integer(string='Año', required=True, default=lambda self: fields.Date.context_today(self).year)
    month = fields.Selection([
        ('01', 'Enero'), ('02', 'Febrero'), ('03', 'Marzo'), ('04', 'Abril'),
        ('05', 'Mayo'), ('06', 'Junio'), ('07', 'Julio'), ('08', 'Agosto'),
        ('09', 'Septiembre'), ('10', 'Octubre'), ('11', 'Noviembre'), ('12', 'Diciembre')
    ], string='Mes', required=True)

    periodo = fields.Char(
        string='Periodo',
        readonly=True,
        compute='_compute_periodo',
        store=True,
    )

    tipo_libro = fields.Selection([
        ('consumidor', 'Consumidor Final'),
        ('credito', 'Crédito Fiscal'),
    ], string='Tipo de Libro', required=True, readonly=True)

    _inherit = ['mail.thread', 'mail.activity.mixin']

    state = fields.Selection([
        ('draft', 'Borrador'),
        ('validated', 'Validado'),
    ], string='Estado', default='draft', tracking=True)

    comentarios = fields.Text(string='Comentarios')

    def action_rectify(self):
        """Abre el wizard para rectificar el libro."""
        self.ensure_one()
        return {
            'name': 'Rectificar Libro',
            'type': 'ir.actions.act_window',
            'res_model': 'libro.rectify.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'active_id': self.id, 'active_model': self._name},
        }

    def rectify_book(self, reason):
        """Método llamado por el wizard para ejecutar la rectificación."""
        self.ensure_one()
        self.message_post(body=f"Libro rectificado. Motivo: {reason}", subtype_xmlid="mail.mt_note")
        self.state = 'draft'

    invoice_line_ids = fields.One2many(
        'libro.ventas.line',
        'periodo_id',
        string='Detalle Ventas',
        domain=[('move_id.state', '=', 'posted')]
    )

    invoice_line_ids_cancelled = fields.One2many(
        'libro.ventas.line',
        'periodo_id',
        string='Detalle Anuladas',
        domain=[('move_id.state', '=', 'cancel')]
    )

    company_currency_id = fields.Many2one(
        'res.currency',
        related='company_id.currency_id',
        readonly=True,
        store=True,
    )

    # Campos calculados
    total_ventas_exentas = fields.Monetary(
        string="Total Ventas Exentas",
        compute="_compute_totales",
        currency_field="company_currency_id",
    )

    total_ventas_gravadas = fields.Monetary(
        string="Total Ventas Gravadas",
        compute="_compute_totales",
        currency_field="company_currency_id",
    )

    total_debito_fiscal = fields.Monetary(
        string="Total Débito Fiscal",
        compute="_compute_totales",
        currency_field="company_currency_id",
    )

    # ----------------- COMPUTADOS -----------------

    @api.depends('year', 'month')
    def _compute_periodo(self):
        for rec in self:
            if rec.year and rec.month:
                month_names = {
                    '01': 'Enero', '02': 'Febrero', '03': 'Marzo', '04': 'Abril',
                    '05': 'Mayo', '06': 'Junio', '07': 'Julio', '08': 'Agosto',
                    '09': 'Septiembre', '10': 'Octubre', '11': 'Noviembre', '12': 'Diciembre'
                }
                rec.periodo = f"{month_names.get(rec.month, '')} {rec.year}"
            else:
                rec.periodo = ''

    year_display = fields.Char(string='Año (Display)', compute='_compute_year_display', store=False)

    @api.depends('year')
    def _compute_year_display(self):
        for rec in self:
            rec.year_display = str(rec.year) if rec.year else ''

    # ----------------- ACCIONES DE SELECCIÓN -----------------

    def action_select_all(self):
        """Seleccionar todas las líneas."""
        for rec in self:
            rec.invoice_line_ids.write({'select': True})

    def action_unselect_all(self):
        """Deseleccionar todas las líneas."""
        for rec in self:
            rec.invoice_line_ids.write({'select': False})

    @api.depends("invoice_line_ids.ventas_exentas",
                 "invoice_line_ids.ventas_gravadas",
                 "invoice_line_ids.debito_fiscal")
    def _compute_totales(self):
        for rec in self:
            rec.total_ventas_exentas = sum(rec.invoice_line_ids.mapped("ventas_exentas"))
            # Sumar ventas_gravadas (subtotal sin IVA) para mostrar en Odoo
            rec.total_ventas_gravadas = sum(rec.invoice_line_ids.mapped("ventas_gravadas"))
            rec.total_debito_fiscal = sum(rec.invoice_line_ids.mapped("debito_fiscal"))

    # ----------------- RESTRICCIONES -----------------

    def write(self, vals):
        """Bloquea edición si el estado no es 'draft', excepto si se cambia el estado."""
        for rec in self:
            if rec.state != 'draft' and 'state' not in vals:
                raise UserError("Solo puedes modificar libros en estado Borrador.")
        return super().write(vals)

    # ----------------- ACCIONES DE ESTADO -----------------

    def action_mark_done(self):
        """Validar el libro."""
        for rec in self:
            rec.state = 'validated'

    def action_reset_to_draft(self):
        """Volver a borrador."""
        for rec in self:
            rec.state = 'draft'

    # ----------------- LÓGICA DE LIBRO -----------------

    def action_load_invoices(self):
        """Generar Detalle: carga facturas del mes según tipo de libro."""
        self.ensure_one()

        if not self.year or not self.month:
            raise UserError("Debe especificar Año y Mes.")

        # Calcular fechas
        date_from = datetime(self.year, int(self.month), 1).date()
        # Último día del mes
        date_to = (date_from + relativedelta(months=1, days=-1))

        # Determinar las compañías a incluir
        if self.incluir_sucursales:
            # Incluir empresa actual + todas sus sucursales
            company_ids = [self.company_id.id] + self.company_id.child_ids.ids
        else:
            # Solo la empresa actual
            company_ids = [self.company_id.id]

        # Limpiar líneas anteriores (tanto normales como anuladas)
        self.invoice_line_ids.unlink()
        self.invoice_line_ids_cancelled.unlink()

        # Definir tipos de documentos según el libro
        if self.tipo_libro == 'consumidor':
            # Consumidor Final: Solo 01, 02, 10, 11 según manual Hacienda
            # Las notas de crédito/débito NO son válidas en Anexo 2
            allowed_doc_types = ['01', '02', '10', '11']
            # IMPORTANTE: Para consumidor final NO incluir out_refund
            move_types = ['out_invoice']
        else:
            # Crédito Fiscal (03) y Notas de Crédito/Débito relacionadas
            allowed_doc_types = ['03', '05', '06']
            move_types = ['out_invoice', 'out_refund']

        # --- 1. CARGAR FACTURAS VALIDAS (POSTED) ---
        invoices = self.env['account.move'].search([
            ('move_type', 'in', move_types),
            ('state', '=', 'posted'),
            ('invoice_date', '>=', date_from),
            ('invoice_date', '<=', date_to),
            ('company_id', 'in', company_ids),
        ], order='invoice_date asc, name asc, id asc')

        lines_values = []
        sequence = 1

        for inv in invoices:
            # Filtrar por tipo de documento DTE
            doc_type_code = inv.l10n_latam_document_type_id.code
            if doc_type_code not in allowed_doc_types:
                continue

            # Mapeo de campos DTE
            numero_documento = inv.name
            numero_control = inv.tgr_l10n_sv_edi_numero_control or ''
            codigo_generacion = inv.tgr_l10n_sv_edi_codigo_generacion or ''
            sello_recepcion = inv.tgr_l10n_sv_edi_sello_recibido or ''
            tipo_documento = inv.l10n_latam_document_type_id.code or ''



            # Montos: DIFERENCIA ENTRE CONSUMIDOR FINAL Y CRÉDITO FISCAL
            # - Consumidor Final: IVA incluido en precio → reportar monto TOTAL
            # - Crédito Fiscal: IVA separado → reportar solo SUBTOTAL sin IVA
            ventas_exentas = 0.0
            ventas_gravadas = 0.0
            debito_fiscal = 0.0
            
            # Nuevos campos para CSV Hacienda
            ventas_exentas_no_sujetas = 0.0
            ventas_no_sujetas = 0.0
            ventas_gravadas_locales = 0.0
            exportaciones_centroamerica = 0.0
            exportaciones_fuera_centroamerica = 0.0
            exportaciones_servicios = 0.0
            ventas_zonas_francas = 0.0
            ventas_cuenta_terceros = 0.0


            # Determinar si la factura tiene impuestos y si están incluidos en precio
            has_taxes = any(line.tax_ids for line in inv.invoice_line_ids)
            
            if has_taxes:
                # Verificar si ALGÚN impuesto tiene price_include=True (Consumidor Final)
                tax_ids = inv.invoice_line_ids.mapped('tax_ids')
                price_include = any(tax.price_include for tax in tax_ids)
                
                # Para TODOS los casos (consumidor y crédito):
                # ventas_gravadas y ventas_gravadas_locales = SUBTOTAL sin IVA
                ventas_gravadas = inv.amount_untaxed
                ventas_gravadas_locales = inv.amount_untaxed
                debito_fiscal = inv.amount_total - inv.amount_untaxed
            else:
                # Facturas sin impuesto son exentas
                ventas_exentas = inv.amount_untaxed
                ventas_gravadas = 0.0
                ventas_gravadas_locales = 0.0
                debito_fiscal = 0.0
            
            # Si es factura de exportación (11), mover a exportaciones
            if tipo_documento == '11':
                # Por defecto a fuera de CA, usuario puede cambiarlo
                exportaciones_fuera_centroamerica = inv.amount_total if price_include else inv.amount_untaxed
                ventas_gravadas_locales = 0.0
                ventas_gravadas = 0.0
                ventas_exentas = 0.0
                debito_fiscal = 0.0  # Exportaciones no tienen débito fiscal

            # Crear diccionario para la línea
            lines_values.append({
                'periodo_id': self.id,
                'sequence': sequence,
                'move_id': inv.id,
                'partner_id': inv.partner_id.id,
                'invoice_date': inv.invoice_date,
                'numero_documento': numero_documento,
                'numero_control': numero_control,
                'codigo_generacion': codigo_generacion,
                'sello_recepcion': sello_recepcion,
                'tipo_documento': tipo_documento,
                'ventas_exentas': ventas_exentas,
                'ventas_gravadas': ventas_gravadas, # Mantener para compatibilidad
                'debito_fiscal': debito_fiscal,
                'amount_total': inv.amount_total,
                # Nuevos campos
                'ventas_exentas_no_sujetas': ventas_exentas_no_sujetas,
                'ventas_no_sujetas': ventas_no_sujetas,
                'ventas_gravadas_locales': ventas_gravadas_locales,
                'exportaciones_centroamerica': exportaciones_centroamerica,
                'exportaciones_fuera_centroamerica': exportaciones_fuera_centroamerica,
                'exportaciones_servicios': exportaciones_servicios,
                'ventas_zonas_francas': ventas_zonas_francas,
                'ventas_cuenta_terceros': ventas_cuenta_terceros,
                'select': True,  # Auto-seleccionar al cargar
            })
            sequence += 1

        # Crear líneas de facturas válidas
        self.env['libro.ventas.line'].create(lines_values)
        
        # Forzar recálculo de totales (igual que en compras)
        self.invalidate_recordset(['invoice_line_ids'])

        # --- 2. CARGAR FACTURAS ANULADAS (CANCEL) ---
        cancelled_invoices = self.env['account.move'].search([
            ('move_type', 'in', ['out_invoice', 'out_refund']),
            ('state', '=', 'cancel'),
            ('invoice_date', '>=', date_from),
            ('invoice_date', '<=', date_to),
            ('company_id', 'in', company_ids),
        ])

        cancelled_lines_values = []
        # Reiniciar secuencia o continuar? Generalmente anuladas tienen su propia lista o se mezclan.
        # Aquí las pondremos en su propia pestaña, así que secuencia propia.
        seq_cancelled = 1

        for inv in cancelled_invoices:
            # Filtrar por tipo de documento DTE
            doc_type_code = inv.l10n_latam_document_type_id.code
            if doc_type_code not in allowed_doc_types:
                continue

            # Mapeo de campos DTE
            numero_documento = inv.name
            numero_control = inv.tgr_l10n_sv_edi_numero_control or ''
            codigo_generacion = inv.tgr_l10n_sv_edi_codigo_generacion or ''
            sello_recepcion = inv.tgr_l10n_sv_edi_sello_recibido or ''
            tipo_documento = inv.l10n_latam_document_type_id.code or ''

            # Para anuladas, los montos suelen ser 0 o se muestran informativamente.
            # El usuario pidió "el mismo filtro", asumiremos que quiere ver los datos aunque estén anuladas.
            # Pero contablemente no suman. En el reporte se verá.
            
            ventas_exentas = 0.0
            ventas_gravadas = 0.0
            debito_fiscal = 0.0

            for line in inv.invoice_line_ids:
                amount_line = line.price_subtotal
                if line.tax_ids:
                    ventas_gravadas += amount_line
                    debito_fiscal += line.price_total - amount_line
                else:
                    ventas_exentas += amount_line

            cancelled_lines_values.append({
                'periodo_id': self.id,
                'sequence': seq_cancelled,
                'move_id': inv.id,
                'partner_id': inv.partner_id.id,
                'invoice_date': inv.invoice_date,
                'numero_documento': numero_documento,
                'numero_control': numero_control,
                'codigo_generacion': codigo_generacion,
                'sello_recepcion': sello_recepcion,
                'tipo_documento': tipo_documento,
                'ventas_exentas': ventas_exentas,
                'ventas_gravadas': ventas_gravadas,
                'debito_fiscal': debito_fiscal,
                'amount_total': inv.amount_total,
            })
            seq_cancelled += 1

        # Crear líneas de facturas anuladas
        self.env['libro.ventas.line'].create(cancelled_lines_values)

    def action_generate_excel(self):
        """Generar archivo Excel (.xlsx) con las facturas seleccionadas."""
        selected = self.invoice_line_ids.filtered(lambda l: l.select)
        if not selected:
            raise UserError("Debe seleccionar al menos una factura.")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise UserError("La librería 'openpyxl' no está instalada. Instálela con: pip install openpyxl")

        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro de Ventas"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Encabezados
        headers = [
            'No', 'Fecha Emisión', 'Número de Documento', 'Número de Control',
            'Código Generación', 'Sello Recepción', 'Cliente', 'Ventas Exentas',
            'Ventas Gravadas', 'Débito Fiscal', 'Total'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Datos
        for row, line in enumerate(selected, start=2):
            ws.cell(row=row, column=1, value=line.sequence or '')
            ws.cell(row=row, column=2, value=str(line.invoice_date) if line.invoice_date else '')
            ws.cell(row=row, column=3, value=line.numero_documento or '')
            ws.cell(row=row, column=4, value=line.numero_control or '')
            ws.cell(row=row, column=5, value=line.codigo_generacion or '')
            ws.cell(row=row, column=6, value=line.sello_recepcion or '')
            ws.cell(row=row, column=7, value=line.partner_id.name or '')
            ws.cell(row=row, column=8, value=line.ventas_exentas or 0)
            ws.cell(row=row, column=9, value=line.ventas_gravadas or 0)
            ws.cell(row=row, column=10, value=line.debito_fiscal or 0)
            ws.cell(row=row, column=11, value=line.amount_total or 0)

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        excel_data = base64.b64encode(output.getvalue())
        output.close()

        tipo_nombre = 'Consumidor_Final' if self.tipo_libro == 'consumidor' else 'Credito_Fiscal'
        attachment = self.env['ir.attachment'].create({
            'name': f'Libro_Ventas_{tipo_nombre}_{self.periodo or ""}.xlsx',
            'type': 'binary',
            'datas': excel_data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_generate_csv(self):
        """Generar CSV con las facturas seleccionadas."""
        # Si es Consumidor Final, usar formato Hacienda (Anexo 2)
        if self.tipo_libro == 'consumidor':
            return self.action_generate_csv_consumidor()
            
        # CRÉDITO FISCAL: Formato Hacienda Anexo 1 (20 columnas A-T)
        selected = self.invoice_line_ids.filtered(lambda l: l.select)
        if not selected:
            raise UserError("Debe seleccionar al menos una factura.")

        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')  # Separador punto y coma

        # SIN encabezados según manual oficial
        
        # Generar filas de datos (20 columnas: A-T)
        for line in selected:
            # A. Fecha Emisión (DD/MM/YYYY)
            fecha_str = line.invoice_date.strftime('%d/%m/%Y') if line.invoice_date else ''
            
            # B. Clase de Documento (1=Impreso, 4=DTE)
            clase_doc = '4' if line.codigo_generacion else '1'
            
            # C. Tipo de Documento (03=CCF, 05=NC, 06=ND)
            tipo_doc = line.tipo_documento or '03'
            
            # D. Número de Resolución (para DTE: número de control sin guiones)
            if line.codigo_generacion:
                # DTE: número de control sin guiones
                numero_resolucion = (line.numero_control or '').replace('-', '')
            else:
                # Impreso: número de resolución real
                numero_resolucion = 'N/A'  # Ajustar según tu sistema
            
            # E. Número de Serie (para DTE: sello de recepción)
            if line.codigo_generacion:
                numero_serie = line.sello_recepcion or ''
            else:
                numero_serie = 'SERIE'  # Ajustar según tu sistema
            
            # F. Número de Documento (para DTE: código de generación sin guiones)
            if line.codigo_generacion:
                numero_documento = line.codigo_generacion.replace('-', '')
            else:
                numero_documento = line.numero_documento or ''
            
            # G. Número de Control Interno (para DTE: dejar en blanco)
            if line.codigo_generacion:
                control_interno = ''
            else:
                control_interno = line.numero_control or line.numero_documento or ''
            
            # H. NIT o NRC del Cliente (sin guiones)
            # Obtener NIT o NRC del partner
            nit_nrc = ''
            if line.partner_id:
                # Buscar VAT (NIT) del partner
                if line.partner_id.vat:
                    nit_nrc = line.partner_id.vat.replace('-', '').replace('/', '')
            
            # I. Nombre del Cliente
            nombre_cliente = line.partner_id.name or ''
            
            # J. Ventas Exentas
            ventas_exentas = f"{line.ventas_exentas:.2f}"
            
            # K. Ventas No Sujetas
            ventas_no_sujetas = "0.00"  # Ajustar si tienes este campo
            
            # L. Ventas Gravadas Locales
            ventas_gravadas = f"{line.ventas_gravadas:.2f}"
            
            # M. Débito Fiscal
            debito_fiscal = f"{line.debito_fiscal:.2f}"
            
            # N. Ventas a Cuenta de Terceros
            ventas_terceros = "0.00"  # Ajustar si tienes este campo
            
            # O. Débito Fiscal por Venta a Terceros
            debito_terceros = "0.00"  # Ajustar si tienes este campo
            
            # P. Total Ventas
            total_ventas = f"{line.amount_total:.2f}"
            
            # Q. DUI del Cliente (9 dígitos, opcional)
            dui_cliente = ''  # Ajustar si tienes este campo en partner
            
            # R. Tipo de Operación (Renta) - desde Enero 2025
            tipo_operacion = line.tipo_operacion_renta or '1'
            
            # S. Tipo de Ingreso (Renta) - desde Enero 2025
            tipo_ingreso = line.tipo_ingreso_renta or '3'
            
            # T. Número de Anexo (siempre 1 para crédito fiscal)
            numero_anexo = '1'
            
            writer.writerow([
                fecha_str,          # A
                clase_doc,          # B
                tipo_doc,           # C
                numero_resolucion,  # D
                numero_serie,       # E
                numero_documento,   # F
                control_interno,    # G
                nit_nrc,            # H
                nombre_cliente,     # I
                ventas_exentas,     # J
                ventas_no_sujetas,  # K
                ventas_gravadas,    # L
                debito_fiscal,      # M
                ventas_terceros,    # N
                debito_terceros,    # O
                total_ventas,       # P
                dui_cliente,        # Q
                tipo_operacion,     # R
                tipo_ingreso,       # S
                numero_anexo,       # T
            ])

        data = base64.b64encode(output.getvalue().encode('utf-8'))
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': f'Libro_Ventas_Credito_Fiscal_Hacienda_{self.periodo or ""}.csv',
            'type': 'binary',
            'datas': data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'text/csv'
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_generate_csv_consumidor(self):
        """Generar CSV formato oficial Hacienda (Anexo 2 - Consumidor Final)."""
        # Para Consumidor Final, incluir TODAS las líneas del periodo
        # (no depender del campo 'select' que solo afecta las líneas visibles en la vista)
        selected = self.invoice_line_ids
        if not selected:
            raise UserError("No hay facturas para exportar. Genere el detalle primero.")

        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')  # Separador punto y coma
        
        # SIN ENCABEZADOS según requerimiento (el ejemplo los muestra pero dice "no deben contener encabezados")
        # El usuario dijo "ejemplo csv" y mostró datos sin encabezados.
        
        # Procesar filas - CADA FACTURA ES UNA LÍNEA INDIVIDUAL
        # No agrupar por fecha, cada DTE tiene su código de generación único
        rows = []
        
        # Procesar todas las líneas individualmente
        for line in selected:
            # Si es DTE (tiene código de generación)
            if line.codigo_generacion:
                # Clase 4 = DTE
                clase = '4'
                # Para DTEs: columnas H e I son el código de generación (Del y Al son iguales para una factura individual)
                doc_del = line.codigo_generacion
                doc_al = line.codigo_generacion
                # Columnas D-G son N/A para DTEs
                resolucion = 'N/A'
                serie = 'N/A'
                control_del = 'N/A'
                control_al = 'N/A'
            else:
                # Clase 1 = Impreso (puede ser 2 = Formulario según tipo)
                clase = '1'
                # Para documentos impresos: columnas H e I son el número de documento
                doc_del = line.numero_documento or ''
                doc_al = line.numero_documento or ''
                # Columnas D-G son la resolución, serie y número de control
                resolucion = 'RESOLUCION'  # TODO: agregar campo en el modelo si es necesario
                serie = 'SERIE'            # TODO: agregar campo en el modelo si es necesario
                control_del = line.numero_control or ''
                control_al = line.numero_control or ''
            
            # Calcular total como suma de columnas
            total_calculado = (line.ventas_exentas + line.ventas_exentas_no_sujetas + 
                             line.ventas_no_sujetas + line.ventas_gravadas_locales + 
                             line.exportaciones_centroamerica + line.exportaciones_fuera_centroamerica + 
                             line.exportaciones_servicios + line.ventas_zonas_francas + 
                             line.ventas_cuenta_terceros)
            
            row = [
                line.invoice_date.strftime('%d/%m/%Y'), # A. Fecha
                clase,                                  # B. Clase (4=DTE, 1=Impreso)
                line.tipo_documento or '01',            # C. Tipo de Documento
                resolucion,                             # D. Resolución
                serie,                                  # E. Serie
                control_del,                            # F. Control Del
                control_al,                             # G. Control Al
                doc_del,                                # H. Doc Del (Código Gen o Num Doc)
                doc_al,                                 # I. Doc Al (Código Gen o Num Doc)
                '',                                     # J. Máquina (vacío)
                f"{line.ventas_exentas:.2f}",           # K
                f"{line.ventas_exentas_no_sujetas:.2f}",# L
                f"{line.ventas_no_sujetas:.2f}",        # M
                f"{line.amount_total:.2f}",             # N (Total con IVA para consumidor)
                f"{line.exportaciones_centroamerica:.2f}", # O
                f"{line.exportaciones_fuera_centroamerica:.2f}", # P
                f"{line.exportaciones_servicios:.2f}",  # Q
                f"{line.ventas_zonas_francas:.2f}",     # R
                f"{line.ventas_cuenta_terceros:.2f}",   # S
                f"{total_calculado:.2f}",               # T (calculado, no amount_total)
                line.tipo_operacion_renta or '1',       # U
                line.tipo_ingreso_renta or '3',         # V
                '2'                                     # W. Anexo (2)
            ]
            rows.append(row)

        # Escribir filas
        for row in rows:
            writer.writerow(row)

        data = base64.b64encode(output.getvalue().encode('utf-8'))
        output.close()

        attachment = self.env['ir.attachment'].create({
            'name': f'Libro_Ventas_Consumidor_Hacienda_{self.periodo or ""}.csv',
            'type': 'binary',
            'datas': data,
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'text/csv'
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_print_report(self):
        """Imprimir: genera el PDF del libro."""
        return self.env.ref('libros_fiscales.report_libro_ventas').report_action(self)
